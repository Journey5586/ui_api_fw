# !/usr/bin/python
# -*- coding: UTF-8 -*- 
# 设置utf-8  显示中文
"""
@Create Time: 2019-08-28 15:03
@Author: guozg 
@File：read_excel_data.py
@Description: 
1.
2.
@Modify Time:  
@Modify Description: 
1.
2.
"""
from openpyxl import load_workbook as lw



class ReadExcelData():
    '''
    读取excel sheet表中的数据，返回的数据主要有两种：  \n
    1：返回的结果为list，类型为[{一行的数据(含表头)},{一行的数据(含表头)},……{一行的数据(含表头)}]  \n
    2：返回的结果为dict，类型为{'case id value':{一行的数据(含表头)},'case id value':{},……'case id value':{}}  \n
    '''

    def __init__(self,excelpath,sheetname):
        ''''''
        self.excel_path = excelpath
        self.sheet_name = sheetname
        self.wb = lw(self.excel_path)
        self.ws = self.wb.get_sheet_by_name(self.sheet_name)
        self.rows = self.ws.max_row
        self.cols = self.ws.max_column

    def get_sheet_data_list(self):
        '''
        返回是一个list ，[{一行数据(含表头)},{一行数据(含表头)},{一行数据(含表头)}]  \n
        返回的结果主是用于进行 ddt的 data参数传递。  \n
        :return: list
        '''
        # # 方法一
        # data_list = []
        # for row in range(1,self.rows+1):
        #     row_data={}
        #     for col in range(1,self.cols+1):
        #         if row > 1:
        #             key = self.ws.cell(1,col).value
        #             value = self.ws.cell(row,col).value
        #             row_data[key] = value
        #
        #     if row>1:
        #         data_list.append(row_data)
        #
        # return data_list

        # 方法二
        data_list = []
        # 第一行为表头，直接从第二获取数据即可。
        for row in range(2, self.rows + 1):
            row_data = {}
            for col in range(1, self.cols + 1):

                key = self.ws.cell(1, col).value
                value = self.ws.cell(row, col).value
                row_data[key] = value

            data_list.append(row_data)

        return data_list



    def get_sheet_data_dict(self):
        '''
        sheet表数据，双层dict，{'case id value':{一行数据(含表头)},'case id value':{一行数据(含表头)}}  \n
        返回的结果主要是用于解决数据依赖。  \n
        :return: dict
        '''
        # # 方法一
        # data_dict = {}
        # for row in range(1,self.rows+1):
        #     key_value = None
        #     row_dict = {}
        #     for col in range(1,self.cols):
        #         if row > 1:
        #             if col == 1:
        #                 key_value = self.ws.cell(row,col).value
        #
        #             key = self.ws.cell(1,col).value
        #             value = self.ws.cell(row,col).value
        #             row_dict[key] = value
        #     if row>1:
        #         data_dict[key_value] = row_dict
        # return data_dict

        # 方法二
        data_dict = {}
        for row in range(2, self.rows + 1):
            key_value = None
            row_dict = {}
            for col in range(1, self.cols):
                if col == 1:
                    key_value = self.ws.cell(row, col).value

                key = self.ws.cell(1, col).value
                value = self.ws.cell(row, col).value
                row_dict[key] = value

            data_dict[key_value] = row_dict

        return data_dict

# help(ReadExcelData)